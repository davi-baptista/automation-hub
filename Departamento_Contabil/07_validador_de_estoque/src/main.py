from openpyxl import load_workbook
import os

from api import api_main


def update_excel_balance(company_name, excel_file, data, fiscal_year):
    try:
        workbook = load_workbook(filename=excel_file)
        
        bd_sheet = workbook['BD']
        
        first_sheet = workbook[workbook.sheetnames[0]]
        
        for row in range(1, bd_sheet.max_row + 1):
            sum_values = 0
            
            for col in range(2, bd_sheet.max_column + 1):
                cell = bd_sheet.cell(row=row, column=col)
                
                cell_name_excel = str(cell.value)
                if cell_name_excel == 'None':
                    continue
                
                cell_name_excel = cell_name_excel.replace('.', '')
                
                for item in data:
                    if item['CODIGO'] == cell_name_excel:
                        sum_values += float(item['SALDO_FINAL'])
                        continue
                    
            if sum_values != 0:
                first_sheet[f'B{row}'].value = sum_values
        
        first_sheet['A20'].value = f'EMPRESA: {company_name}'
        first_sheet['A21'].value = f'ANO EXERCICIO: {fiscal_year}'
        
        output_path = os.path.dirname(excel_file)
        final_excel_file = os.path.join(output_path, 'validacao_estoque.xlsx')
        workbook.save(filename=final_excel_file)
        return 'Excel file saved successfully'
    
    except Exception as e:
        return f'Failed to save the file: {e}'

def main(cnpj, start_date, end_date):
    company_name, excel_path, data, status = api_main(cnpj, start_date, end_date)
    
    fiscal_year = start_date.split('-')[-1]
    
    if status == 'Found company data':
        excel_status = update_excel_balance(company_name, excel_path, data, fiscal_year)
        return excel_path, excel_status
    else:
        return None, status

if __name__ == '__main__':
    status_main = main('13311185000105', '01-01-2023', '12-31-2023')
    print(status_main)