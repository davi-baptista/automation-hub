import sys
import pandas as pd
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit, QTextEdit, QFileDialog, QMessageBox
from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtCore import Qt
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers

class MainWindow(QWidget):
    
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Balancete Excel para Padrão Accountify")
        self.setGeometry(100, 100, 600, 400)

        self.dark_mode = True
        self.setup_ui()
        self.apply_theme()
        
        
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        print('Interface aberta')

        # Input layout for the CNPJ input and buttons
        input_layout = QHBoxLayout()
        layout.addLayout(input_layout)

        self.company_cnpj_input = QLineEdit()
        self.company_cnpj_input.setPlaceholderText("Coloque aqui o diretorio com os arquivos excel")
        input_layout.addWidget(self.company_cnpj_input)

        self.choose_button = QPushButton("Buscar")
        self.choose_button.clicked.connect(self.choose_directory)
        input_layout.addWidget(self.choose_button)
        
        self.send_button = QPushButton("Enviar")
        self.send_button.clicked.connect(self.send_directory)
        input_layout.addWidget(self.send_button)

        self.toggle_theme_button = QPushButton("Alterar Tema")
        self.toggle_theme_button.clicked.connect(self.toggle_theme)
        layout.addWidget(self.toggle_theme_button)

        self.log_area = QTextEdit(readOnly=True)
        layout.addWidget(self.log_area)

    
    def toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self.apply_theme()


    def apply_theme(self):
        if self.dark_mode:
            apply_dark_mode(app)
        else:
            apply_light_mode(app)
            
    
    def log(self, message):
        self.log_area.append(message)
        QApplication.processEvents()
        
        
    def send_directory(self):
        directory = self.company_cnpj_input.text()
        
        if directory:
            self.log(f'Diretório selecionado: {directory}.')
        
            if not os.path.exists(directory):
                self.log(f'Verifique se o caminho do arquivo esta correto.\n')
                QMessageBox.warning(self, "Erro", 'Caminho invalido! Procure a pasta novamente.')
                return
            
            self.log(f'Iniciando processamento...')
            
            os.makedirs(os.path.join(directory, 'accountify'), exist_ok=True)
            for file in os.listdir(directory):
                if file.endswith('.xls'):
                    self.log(f'Processando excel: {os.path.basename(file)}')
                    file_path = os.path.join(directory, file)
                    destiny_file_path = os.path.join(directory, 'accountify', file.replace('.xls', '_modificado.xlsx'))
                    
                    df_modificado = self.processar_arquivo(file_path)
                    df_modificado.to_excel(destiny_file_path, index=False, header=False)
                    
                    wb = load_workbook(destiny_file_path)
                    planilha = wb.active
                    for linha in planilha.iter_rows(min_col=3, max_col=3):
                        for celula in linha:
                            if celula.value:
                                valor_formatado = float(str(celula.value).replace('.', '').replace(',', '.'))
                                celula.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                                celula.value = valor_formatado
                    
                    wb.save(destiny_file_path)
                    
                if file.endswith('.xlsx'):
                    self.log(f'Processando excel: {os.path.basename(file)}')
                    file_path = os.path.join(directory, file)
                    destiny_file_path = os.path.join(directory, 'accountify', file.replace('.xlsx', '_modificado.xlsx'))
                    
                    df_modificado = self.processar_arquivo(file_path)
                    df_modificado.to_excel(destiny_file_path, index=False, header=False)
                    
            self.log(f'Programa finalizado e arquivos salvos na pasta "accountify"')
            
            
    def processar_arquivo(self, caminho_arquivo):
        # Ler o arquivo Excel
        df = pd.read_excel(caminho_arquivo, sheet_name="Report", header=None)

        # Separar a primeira coluna pelo hífen
        df[[0, 1]] = df[0].str.split('-', n=1, expand=True)

        # Separar a coluna E por letra
        df[[4, 5]] = df[4].str.extract(r'([\d\.,]+)\s*(\D*)')

        # Excluir linhas vazias
        df = df.dropna()

        # Excluir todas as colunas exceto A e E
        colunas_para_manter = [0, 1, 4, 5]
        df = df[colunas_para_manter]
        df = df[df[4] != "0,00"]

        return df
    
    
    def choose_directory(self):
        directory = QFileDialog.getExistingDirectory(self, "Escolher diretório")
        if directory:
            self.company_cnpj_input.setText(directory)


def apply_dark_mode(app):
    app.setStyle("Fusion")
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor(53, 53, 53))
    palette.setColor(QPalette.WindowText, Qt.white)
    palette.setColor(QPalette.Base, QColor(25, 25, 25))
    palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
    palette.setColor(QPalette.ToolTipBase, Qt.white)
    palette.setColor(QPalette.ToolTipText, Qt.white)
    palette.setColor(QPalette.Text, Qt.white)
    palette.setColor(QPalette.Button, QColor(53, 53, 53))
    palette.setColor(QPalette.ButtonText, Qt.white)
    palette.setColor(QPalette.BrightText, Qt.red)
    palette.setColor(QPalette.Link, QColor(42, 130, 218))
    palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
    palette.setColor(QPalette.HighlightedText, Qt.black)
    app.setPalette(palette)


def apply_light_mode(app):
    app.setPalette(app.style().standardPalette())


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    print('Abrindo interface')
    window.show()
    sys.exit(app.exec_())