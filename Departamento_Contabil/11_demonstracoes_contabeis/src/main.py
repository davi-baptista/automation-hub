from openpyxl import load_workbook
import os
import re
from datetime import date

from api import api_main


def formatar_cnpj(cnpj):
    # Remove qualquer caracter que não seja dígito
    cnpj = re.sub(r'\D', '', cnpj)

    # Aplica a formatação padrão do CNPJ
    cnpj_formatado = re.sub(r"(\d{2})(\d{3})(\d{3})(\d{4})(\d{2})", r"\1.\2.\3/\4-\5", cnpj)

    return cnpj_formatado


def update_excel_balance(company_name, cnpj, year, column, excel_file, data):
    try:
        workbook = load_workbook(filename=excel_file)
        
        bd_sheets = ['ATIVO - BD', 'PASSIVO E PL - BD', 'DRE - BD']
        main_sheets = ['ATIVO', 'PASSIVO E PL', 'DRE']
        
        for bd_sheet_name, main_sheet_name in zip(bd_sheets, main_sheets):
            bd_sheet = workbook[bd_sheet_name]
            main_sheet = workbook[main_sheet_name]
            
            for row in range(1, bd_sheet.max_row + 1):
                sum_values = 0
                
                for col in range(2, bd_sheet.max_column + 1):
                    cell = bd_sheet.cell(row=row, column=col)
                    
                    cell_name_excel = str(cell.value)
                    if cell_name_excel == 'None':
                        continue
                    
                    cell_name_excel = cell_name_excel.replace('.', '')
                    
                    for item in data:
                        if item['CODIGO'] == cell_name_excel:
                            sum_values += float(item['SALDO_FINAL'])
                            continue
                        
                if sum_values != 0:
                    main_sheet[f'{column}{row}'].value = sum_values
            
            if main_sheet_name == 'ATIVO':
                main_sheet['A42'].value = f'Fortaleza(CE), 31 de dezembro de {year}'
                main_sheet['A10'].value = f'BALANÇO PATRIMONIAL DOS EXERCÍCIOS FINDOS EM 31 DE DEZEMBRO DE {year} E {year-1}'
            elif main_sheet_name == 'PASSIVO E PL':
                main_sheet['A53'].value = f'Fortaleza(CE), 31 de dezembro de {year}'
            elif main_sheet_name == 'DRE':
                main_sheet['A41'].value = f'Fortaleza(CE), 31 de dezembro de {year}'

            main_sheet[f'{column}16'] = year
            main_sheet['A6'].value = company_name
            main_sheet['A7'].value = cnpj
        
        final_excel_file = excel_file
        if 'demonstracoes' not in os.path.basename(excel_file):
            output_path = os.path.dirname(excel_file)
            final_excel_file = os.path.join(output_path, f'demonstracoes_{year}.xlsx')
            
        workbook.save(filename=final_excel_file)
        return 'Excel file saved successfully', final_excel_file
    
    except Exception as e:
        return f'Failed to save the file: {e}', final_excel_file


def first_last_day_of_year_formatted(year):
    """
    Given a year, returns the first and last day of that year in the format MM-DD-YYYY.
    """
    year = int(year)
    first_day = date(year, 1, 1).strftime("%m-%d-%Y")
    last_day = date(year, 12, 31).strftime("%m-%d-%Y")
    return first_day, last_day


def main(cnpj, year):
    start_date, end_date = first_last_day_of_year_formatted(year)
    company_name, excel_path, data, status = api_main(cnpj, start_date, end_date)
    
    if status == 'Found company data':
        year_anterior = int(year) - 1
        previous_start_date, previous_end_date = first_last_day_of_year_formatted(year_anterior)
        company_name, excel_path, previous_data, previous_status = api_main(cnpj, previous_start_date, previous_end_date)
        if previous_status == 'Found company data':
            cnpj_formatado = formatar_cnpj(cnpj)
            excel_status, final_excel_file = update_excel_balance(company_name, cnpj_formatado, year, 'G', excel_path, data)
            
            year_anterior = int(year) - 1
            previous_excel_status, _ = update_excel_balance(company_name, cnpj_formatado, year_anterior, 'H', final_excel_file, previous_data)
            
            final_status = 'Archive error'
            if excel_status == 'Excel file saved successfully' and previous_excel_status == 'Excel file saved successfully':
                final_status = 'Excel file saved successfully'
            elif excel_status != 'Excel file saved successfully' and previous_excel_status == 'Excel file saved successfully':
                final_status = 'Excel file error in actual year'
            elif excel_status == 'Excel file saved successfully' and previous_excel_status != 'Excel file saved successfully':
                final_status = 'Excel file error in previous year'
            return excel_path, final_status
        else:
            return None, previous_status
    else:
        return None, status

if __name__ == '__main__':
    status_main = main('13311185000105', 2023)
    print(status_main)