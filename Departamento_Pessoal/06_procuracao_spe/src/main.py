from botcity.core import DesktopBot
import pyautogui
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font

from procuracao_spe import main, fechar_processo

class Bot(DesktopBot):
    
    
    def action(self, execution=None):
        pasta_certificados = r'c:\Users\davi.inov\Desktop\CERTIFICADOS NOVOS'
        caminho_excel_retorno = r'C:\Users\davi.inov\Desktop\Projetos\Depa  rtamento_Pessoal\06_procuracao_spe\data\relatorio.xlsx'
        senhas_fixas = ['1234', '123456', '12345678']
        
        # Primeiro, mapeia todos os arquivos .txt relevantes em cada diretório
        txt_map = self.mapear_arquivos_txt(pasta_certificados)
        
        if os.path.exists(caminho_excel_retorno):
            df = pd.read_excel(caminho_excel_retorno)
        else:
            df = pd.DataFrame(columns=['Nome do Certificado', 'Status da Procuração', 'Caminho do certificado', 'Caminho da procuracao'])
            
        # Percorre o diretório e subdiretórios em busca de certificados
        for root, dirs, files in os.walk(pasta_certificados):
            # Verifica se 'vencido' não faz parte do caminho atual
            if 'vencido' in root.lower():
                continue
            
            for file in files:
                try:
                    nome_certificado = 'Desconhecido'  # Valor padrão
                    status_procuracao = 'Erro desconhecido'  # Valor padrão
                    caminho_procuracao = ''
                    caminho_completo = 'Deconhecido'  # Valor padrão
                    
                    # Verifica se o arquivo tem uma das extensões desejadas (.p12, .pfx)
                    if file.endswith(('.p12', '.pfx')):
                        caminho_completo = os.path.join(root, file)
                        nome_certificado, _ = os.path.splitext(file)
                        
                        # Verifica se o caminho_completo ja existe no excel, se ja existir, pula pro proximo
                        if df['Caminho do certificado'].isin([caminho_completo]).any():
                            continue
                            
                        # Chama capturar_e_abrir_arquivo para processar o certificado encontrado
                        status = self.capturar_e_abrir_arquivo(root, file, caminho_completo, senhas_fixas, txt_map)
                        status_procuracao = status
                        print(status)
                        if status != 'Nao achou nenhuma senha':
                            self.abrir_certmgr()
                            status_certificado = self.verificar_validade()
                            status_procuracao = status_certificado
                            nome_certificado = self.pegar_certificado()
                            print(nome_certificado)
                            print(status_certificado)
                            
                            if status_certificado == 'Certificado valido':
                                status_procuracao = 'Processo de procuracao nao iniciado'
                                status_procuracao, caminho_procuracao = main(nome_certificado)
                            
                            self.excluir_certificado()
                            self.wait(500)
                            self.alt_f4()
                            
                        df = self.adicionar_ao_excel(df, nome_certificado, status_procuracao, caminho_completo, caminho_procuracao, caminho_excel_retorno)
                
                except Exception as e:
                    print(f'Erro nos certificados {e}')
                    if status_procuracao == 'Erro desconhecido':
                        self.resetar_certificados_e_fechar()
                        continue
                    
                    df = self.adicionar_ao_excel(df, nome_certificado, status_procuracao, caminho_completo, caminho_procuracao, caminho_excel_retorno)
                    self.resetar_certificados_e_fechar()
                    continue
        
        print('Codigo terminou de rodar completamente.')
        self.adicionar_hyperlink(df, caminho_excel_retorno)
        
        
    def mapear_arquivos_txt(self, pasta):
        txt_map = {}
        for root, dirs, files in os.walk(pasta):
            if 'vencido' in root.lower():
                continue
            for file in files:
                if file.endswith('.txt') or not re.match(r'.*\.[a-zA-Z]{3}$', file):
                    if root not in txt_map:
                        txt_map[root] = []
                    txt_map[root].append(file)
        return txt_map
        
        
    def capturar_e_abrir_arquivo(self, root, file, caminho_completo, senhas_fixas, txt_map):
        self.execute(caminho_completo)
        if not self.find( "certificados", matching=0.97, waiting_time=30000):
            self.not_found("certificados")
        self.enter()
        self.enter()
        
        achou = None
        for senha in senhas_fixas:
            self.control_a()
            self.paste(senha)
            self.enter()
            if self.find( "senha_incorreta", matching=0.97, waiting_time=1000):
                self.enter()
            else:
                achou = f'Senha {senha}'
                self.enter()
                self.enter()
                if not self.find( "importacao_exito", matching=0.97, waiting_time=30000):
                    self.not_found("importacao_exito")
                self.enter()
                return achou
        
        # Remove a extensão do arquivo antes de chamar search_password
        nome_arquivo_sem_extensao, _ = os.path.splitext(file)
        senhas_potenciais = self.search_password(nome_arquivo_sem_extensao, senhas_fixas)
        
        # Usa o mapeamento de arquivos .txt para buscar mais senhas
        if root in txt_map:
            for file_txt in txt_map[root]:
                nome_txt_sem_extensao, _ = os.path.splitext(file_txt)
                senhas_potenciais.extend(self.search_password(nome_txt_sem_extensao, senhas_fixas))
                
        # Remove possíveis duplicatas mantendo a ordem
        senhas_potenciais = list(dict.fromkeys(senhas_potenciais))
        
        print(f"Certificado: {file}, Senhas Potenciais: {senhas_potenciais}")
        
        achou = None
        for senha in senhas_potenciais:
            self.control_a()
            self.paste(senha)
            self.enter()
            if self.find( "senha_incorreta", matching=0.97, waiting_time=1000):
                self.enter()
            else:
                achou = f'Senha {senha}'
                self.enter()
                self.enter()
                if not self.find( "importacao_exito", matching=0.97, waiting_time=30000):
                    self.not_found("importacao_exito")
                self.enter()
                return achou
        
        self.alt_f4()
        return 'Nao achou nenhuma senha'
        
        
    def abrir_certmgr(self):
        self.execute(r'C:\Windows\System32\certmgr.msc')
        
        if not self.find( "certmgr", matching=0.97, waiting_time=30000):
            self.not_found("certmgr")
        self.type_right()
        self.type_right()
        self.type_right()
        
        if not self.find( "emitido_para", matching=0.97, waiting_time=30000):
            self.not_found("emitido_para")
        self.tab()
    
    
    def verificar_validade(self):
        pyautogui.hotkey('alt', 'o', 'a')
        if not self.find( "informacoes_certificado", matching=0.97, waiting_time=30000):
            self.not_found("informacoes_certificado")
            
        status_procuracao = 'Certificado invalido'
        if not self.find( "expirou_ou_nao_valido", matching=0.97, waiting_time=1000):
            status_procuracao = 'Certificado valido'
        self.alt_f4()
        return status_procuracao
    
    
    def excluir_certificado(self):
        self.delete()
        if not self.find( "excluir_certificado", matching=0.97, waiting_time=5000):
            self.not_found("excluir_certificado")
        self.enter()
        
        
    def search_password(self, texto, senhas_fixas):
        # Inicializando o dicionário de resultados
        resultado_senha = []
        
        # Busca por padrões que incluam "senha", seguidos por caracteres como : - ou espaços, e então captura a senha
        padrao = re.compile(r'senha[s]?[\s:_-]*\s*(\S+)|(\S+)\s*senha[s]?[\s:_-]*\s*(\S+)', re.IGNORECASE)
        resultados = padrao.findall(texto)
        
        for resultado in resultados:
            for possivel_senha in filter(None, resultado):
                if possivel_senha not in senhas_fixas and possivel_senha not in resultado_senha:
                    resultado_senha.append(possivel_senha)
                
        # Capturando a primeira e a última palavra do texto
        palavras = texto.split()
        if palavras:
            primeira_palavra, ultima_palavra = palavras[0], palavras[-1]
            if primeira_palavra not in senhas_fixas:
                resultado_senha.append(primeira_palavra)
            if ultima_palavra not in senhas_fixas and ultima_palavra != primeira_palavra:
                resultado_senha.append(ultima_palavra)
                
        return resultado_senha
    
    
    def pegar_certificado(self):
        pyautogui.hotkey('alt', 'o', 'a')
        self.tab()
        self.type_right()
        self.tab()
        self.tab()
        self.type_down()
        self.type_down()
        self.type_down()
        self.type_down()
        self.type_down()
        self.type_down()
        self.type_down()
        self.tab()
        self.type_up()
        pyautogui.hotkey('ctrl', 'right')
        pyautogui.hotkey('ctrl', 'right')
        self.hold_shift()
        self.key_end()
        self.type_left()
        self.release_shift()
        self.control_c()
        self.alt_f4()
        
        buffer = self.get_clipboard()
        return buffer
        
        
    def adicionar_ao_excel(self, df, nome_certificado, status_procuracao, caminho_certificado, caminho_procuracao, caminho_excel):
        # Cria um DataFrame para a nova linha
        nova_linha_df = pd.DataFrame({
            'Nome do Certificado': [nome_certificado],
            'Status da Procuração': [status_procuracao],
            'Caminho do certificado': [caminho_certificado],
            'Caminho da procuracao': [caminho_procuracao]
        })
        
        # Usa pd.concat para adicionar a nova linha ao DataFrame existente
        df = pd.concat([df, nova_linha_df], ignore_index=True)

        # Salva o DataFrame atualizado em um arquivo Excel
        df.to_excel(caminho_excel, index=False)
        return df
        
        
    def resetar_certificados_e_fechar(self):
        self.abrir_certmgr()
        self.excluir_certificado()
        fechar_processo('mmc.exe')
        
        
    def adicionar_hyperlink(df, caminho_excel):
        wb = load_workbook(caminho_excel)
        ws = wb.active

        for row, caminho in enumerate(df['Caminho da procuracao'], start=2):
            if caminho and os.path.exists(caminho):  # Verifica se o caminho não está vazio e existe
                ws.cell(row=row, column=4).hyperlink = 'file:///' + caminho.replace('\\', '/')
                ws.cell(row=row, column=4).font = Font(color='0000FF', underline='single')
            else:
                print(f"Célula na linha {row}, coluna 4 está vazia ou contém um caminho inválido. Hiperlink não adicionado.")

        wb.save(caminho_excel)
    
    
    def not_found(self, label):
        print(f"Element not found: {label}")
        raise Exception(f'Element not found {label}')


if __name__ == '__main__':
    Bot.main()