from botcity.core import DesktopBot
from botcity.web import WebBot, Browser
import os
import shutil
import pyautogui
from pywinauto import Application
import openpyxl
import datetime
import pandas as pd


class Bot(DesktopBot):
    caminho_saida_erro = r'C:\Users\davi.inov\Desktop\Projetos\Departamento_Pessoal\02. Recolher_FAP\Saidas\Relatorio_Erros\SituacaoFAPs.xlsx'
    pasta_downloads = r'C:\Users\davi.inov\Downloads'
    pasta_destino = r'C:\Users\davi.inov\Desktop\Projetos\Departamento_Pessoal\02. Recolher_FAP\Saidas\FAPs'
    
    
    def add_certificate(self):
        self.enter()
        self.tab()
        self.type_right()
        self.tab()
        self.tab()
        self.type_down()
        self.type_down()
        self.type_down()
        self.type_down()
        self.type_down()
        self.type_down()
        self.type_down()
        self.tab()
        self.type_up()
        pyautogui.hotkey('ctrl', 'right')
        pyautogui.hotkey('ctrl', 'right')
        self.hold_shift()
        self.key_end()
        self.type_left()
        self.release_shift()
        self.control_c()
        self.alt_f4()
        self.type_down()
        
        
    def delete_certificate(self):
        self.type_up()
        pyautogui.press('delete')
        self.wait(200)
        self.enter()
        self.wait(200)
        self.type_down()
        self.type_up()
        
        
    def sanitize_directory_name(self, path):
        # Obtém a última parte do caminho
        base_path, directory_name = os.path.split(path)
        
        # Substitui caracteres inválidos por um espaço em branco
        invalid_characters = r'<>:"/\|?*'
        sanitized_name = ''.join(char if char not in invalid_characters else ' ' for char in directory_name)

        # Junta o caminho base com o novo nome sanitizado
        sanitized_path = os.path.join(base_path, sanitized_name)

        return sanitized_path
    
    
    def recent_download(self, pasta_downloads, pasta_destino, buffer):
        # Lista todos os arquivos na pasta de downloads
        arquivos = [os.path.join(pasta_downloads, arquivo) for arquivo in os.listdir(pasta_downloads)]
        
        # Verifica se tem arquivos encontrados
        if arquivos:
            # Encontra o arquivo mais recente modificado com base no tempo de modificação
            arquivo_mais_recente = max(arquivos, key=os.path.getmtime)
            
            clipboard = self.get_clipboard().split('/')
            ultima_parte = clipboard[-1]
            #-Tirando caracteres invalidos do nome da empresa
            company_name = self.sanitize_directory_name(buffer)
            # Construa o caminho de destino
            caminho_destino = os.path.join(pasta_destino, company_name, ultima_parte, os.path.basename(arquivo_mais_recente))
            # Verifica se a pasta de destino existe, e se não existir, cria-a
            if not os.path.exists(os.path.join(pasta_destino, company_name, ultima_parte)):
                os.makedirs(os.path.join(pasta_destino, company_name, ultima_parte))
            else:
                raise Exception('A pasta já existe')
            # Move o arquivo para a pasta de destino
            shutil.move(arquivo_mais_recente, caminho_destino)

            print(f'O arquivo mais recente foi movido para: {caminho_destino}')
            return caminho_destino
        
        else:
            print('Não foi encontrado nenhum arquivo na pasta de downloads')
            raise Exception(f'Não foi encontrado nenhum arquivo na pasta de downloads')
    
    
    def close_process(self, nome_do_processo):
        app = Application(backend="uia").connect(path='nome_do_processo')
        janela = app.window()
        janela.close()
    
    
    def unexpected_error(self):
        if self.find( "cliqueaqui", matching=0.97, waiting_time=3000):
            self.click()
        if not self.find( "entrargov", matching=0.97, waiting_time=60000):
            self.not_found("entrargov")
        pyautogui.hotkey('alt', 'a')
        self.kb_type('n')
        self.kb_type('acesso.gov.br')
        self.enter()
        if not self.find( "certificadodigital", matching=0.97, waiting_time=30000): #Abre os certificados
            self.not_found("certificadodigital")
        self.click()
        if self.find( "autorizar", matching=0.97, waiting_time=3500):
            self.click()
        if not self.find( "vincularempresas", matching=0.97, waiting_time=30000):
            self.not_found("vincularempresas")
        self.click()
        if self.find( "autorizar2", matching=0.97, waiting_time=3500):
            self.click()
        if not self.find( "vincularempresas2", matching=0.97, waiting_time=30000):
            self.not_found("vincularempresas2")
        self.wait(1500)
        self.enter()
        if not self.find( "vincularempresas2", matching=0.97, waiting_time=30000):
            self.not_found("vincularempresas2")
        self.click()
        self.wait(2000)
        self.enter()
        if not self.find( "vincular", matching=0.97, waiting_time=30000):
            self.not_found("vincular")
        self.wait(1000)
        self.click()
        pyautogui.hotkey('ctrl', 'f4')
        if not self.find( "entrargov", matching=0.97, waiting_time=30000):
            self.not_found("entrargov")
        self.click()
        
        
    def adicionar_ao_excel(caminho_do_excel, dados):
        try:
            # Tenta carregar o arquivo Excel existente
            wb = openpyxl.load_workbook(caminho_do_excel)
            writer = pd.ExcelWriter(caminho_do_excel, engine='openpyxl') 
            writer.book = wb

            # Adiciona os dados ao final da última linha usada
            dados.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet'].max_row)

            # Salva as alterações
            writer.save()

        except FileNotFoundError:
            # Se o arquivo não existir, cria um novo
            dados.to_excel(caminho_do_excel, index=False)


    #-Função que implementa a lógica toda
    def action(self, execution=None):
        workbook = openpyxl.Workbook()  
        sheet = workbook.active
        sheet.append(['Certificado digital', 'Caminho do arquivo', 'Situação'])
        
        self.execute(r'C:\Windows\System32\certmgr.msc')
        if not self.find( "certmgr", matching=0.97, waiting_time=30000):
            self.not_found("certmgr")
        self.type_right()
        self.type_right()
        self.type_right()
        if not self.find( "emitidopara", matching=0.97, waiting_time=30000):
            self.not_found("emitidopara")
        self.tab()
        buffer = None
        comparador = None
        while True:
            try:
                self.add_certificate()
                contador = 1
                if comparador == self.get_clipboard():
                    print('Finalizado')
                    self.alt_f4()
                    break
                buffer = self.get_clipboard()
                self.execute(r'C:\Program Files\Mozilla Firefox\private_browsing.exe')
                if not self.find( "firefoxlogo", matching=0.97, waiting_time=30000):
                    self.not_found("firefoxlogo")
                self.kb_type("https://fap.dataprev.gov.br/")
                self.enter()
                if not self.find( "entrargov", matching=0.97, waiting_time=30000):
                    self.not_found("entrargov")
                self.click()
                if not self.find( "certificadodigital", matching=0.97, waiting_time=30000): #Abre os certificados
                    self.not_found("certificadodigital")
                self.click_relative(x=50, y=0)
                if not self.find( "detalhes_certificado", matching=0.97, waiting_time=30000): #Abre os certificados
                    self.not_found("detalhes_certificado")
                self.kb_type(buffer)
                self.enter()
                self.wait(1000)
                if not self.find( "site_carregado", matching=0.97, waiting_time=60000): #Abre os certificados
                    self.not_found("site_carregado")
                if self.find( "verificacaoduasetapas", matching=0.97, waiting_time=3000):
                    raise Exception('Verificação em duas etapas')
                if self.find( "erroinesperado", matching=0.97, waiting_time=10000):
                    self.unexpected_error()
                if self.find( "termodeuso", matching=0.97, waiting_time=3000):
                    self.click()
                    if self.find( "erroinesperado", matching=0.97, waiting_time=10000):
                        self.unexpected_error()
                if self.find( "autorizar", matching=0.97, waiting_time=3000):
                    self.click()
                    if self.find( "erroinesperado", matching=0.97, waiting_time=10000):
                        self.unexpected_error()
                self.wait(2000)
                if not self.find( "procuracao", matching=0.97, waiting_time=60000):
                    self.not_found("procuracao")
                self.click()
                self.wait(2000)
                if not self.find( "incluir", matching=0.97, waiting_time=30000):
                    self.not_found("incluir")
                self.click()
                if not self.find( "tipoprocuracao", matching=0.97, waiting_time=30000):
                    self.not_found("tipoprocuracao")
                self.wait(500)
                self.double_click()
                self.type_down()
                self.enter()
                self.wait(2000)
                self.tab()
                self.type_down()
                self.enter()
                self.wait(1500)
                self.tab()
                self.kb_type('05.494.895')
                self.tab()
                self.tab()
                data_hora_atual = datetime.datetime.now()
                dia = f'{data_hora_atual.day:02}'
                mes = f'{data_hora_atual.month:02}'
                ano = data_hora_atual.year
                self.kb_type(str(dia))
                self.kb_type(str(mes)) 
                self.kb_type(str(ano))
                self.tab()
                self.tab()
                self.kb_type('3112')
                ano_futuro = ano + 5
                self.kb_type(str(ano_futuro))
                self.tab()
                self.tab()
                self.tab()
                self.enter()
                if self.find( "existeprocuracao", matching=0.97, waiting_time=5000) or self.find( "procuracaorealizada", matching=0.97, waiting_time=5000):
                    if not self.find( "govbrhome", matching=0.97, waiting_time=10000):
                        self.not_found("govbrhome")
                    self.click()
                
                buffer2 = []
                while True:
                    if not self.find( "govbrhome", matching=0.97, waiting_time=10000):
                        self.not_found("govbrhome")
                    self.click()
                    if not self.find( "consulta_fap", matching=0.97, waiting_time=10000):
                        self.not_found("consulta_fap")
                    self.click()
                    self.wait(2000)
                    if not self.find( "vigencia", matching=0.97, waiting_time=10000):
                        self.not_found("vigencia")
                    self.click()
                    self.type_down()
                    self.enter()
                    self.tab()
                    self.type_down()
                    self.enter()
                    if self.find( "existeprocuracao", matching=0.97, waiting_time=5000):
                        raise Exception('Estabelecimento inexistente ou não constante na base de dados do FAP vigência 2024.')
                    self.tab()
                    for i in range(contador):
                        self.type_down()
                    contador += 1
                    self.enter()
                    self.control_a()
                    self.control_c()
                    if self.get_clipboard() in buffer2:
                        break
                    buffer2.append(self.get_clipboard())
                    self.tab()
                    self.enter()
                    if not self.find( "consultarfap", matching=0.97, waiting_time=30000):
                        self.not_found("consultarfap")
                    self.tab()
                    self.enter()
                    if not self.find( "relatorioaberto", matching=0.97, waiting_time=30000):
                        self.not_found("relatorioaberto")
                    caminho_arquivo = self.recent_download(self.pasta_downloads, self.pasta_destino, buffer)
                    pyautogui.hotkey('ctrl', 'f4')
                self.alt_f4()
                sheet.append([buffer, caminho_arquivo, 'Baixado'])
                self.wait(500)
                comparador = self.get_clipboard()
                self.delete_certificate()
            except Exception as e:
                self.alt_f4()
                self.wait(500)
                self.delete_certificate()
                print(f'Erro {e}')
                sheet.append([buffer, '' , f'Erro {e}'])
                continue
            finally:
                workbook.save(self.caminho_saida_erro)
        
        
    def not_found(self, label):
        raise Exception(f'Element not found: {label}')


if __name__ == '__main__':
    Bot.main()